#!/usr/bin/env python
# -*- coding: utf-8 -*-

__author__ = "Raphaël LEBER"

from CalendarSlot import CalendarSlot
import csv
from openpyxl import load_workbook
from datetime import datetime
from datetime import timedelta
from copy import deepcopy
import string
from math import fmod
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
import unittest

class GCalendarCpeGenerator:

    def __init__(self):
        wb = load_workbook(filename='Planning-ROS-S9-20-21_v3.xlsx', data_only=True) # file to open
        self.ws = wb.active

        # Date cell of the first day
        self.start_cell = 'A24'

        # End cell of the last event
        self.end_cell = 'Y84'

        # Legend of the events to consider
        self.legend_range = self.ws['AI25':'AI59']

        # Columns to skip
        self.skip = [] # e.i. 'T'

        # Filter contributor. Leave '' for no filtering
        self.contributor = ''


        for sk in reversed(self.skip):
            self.ws.delete_cols( self.col2num(sk) )

        self.planning_range = self.ws[ self.start_cell:self.end_cell ]

        self.start_date = deepcopy( self.ws[self.start_cell].value )

        self.legend = []
        self.day_content = []

        self.calendarSlots = []



    def getLegend(self):

        # Get legend
        for row in self.legend_range:
            if row[0].value != None:
                self.legend.append(str(row[0].value).split(maxsplit=1)[0])

        print(self.legend)



    def col2num(self, col):
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num



    def generateCalendarSlots(self):
        CS = CalendarSlot(datetime.now())
        current_day_date = self.start_date
        #print('########################################  ' + str(current_day_date))


        for row_i, row in enumerate(self.planning_range) :
            for cell_i, cell in enumerate(row):

                dayOffset = 7*row_i + int(cell_i/5)
                h, m = CS.TimeOfSlot( cell_i % 5)
                #current_day_date = self.start_date + timedelta(days=dayOffset, hours=h, minutes=m)
                if (h, m) == (0, 0): # si on est sur une cellule 'date'
                    if cell.value != None:
                        current_day_date = cell.value 
    
                current_day_with_time = current_day_date + timedelta(hours=h, minutes=m)

                slotText = str(cell.value).replace('\n', ' ').strip()
                SlotTextSplit = slotText.split()

                modules_in_slot = []
                slot_modules = slotText.replace('_x000D_','').split('//') #_x000D_ est un caractère unicode résiduel de office365 en ligne. Eq de `\n`
                for i,s in enumerate(slot_modules) :
                    modules_in_slot.append(s.split()[0].lower())

                ModuleName1 = SlotTextSplit[0]
                #if slot

                #if ModuleName.lower() in [x.lower() for x in self.legend] :
                #if list(set([x.lower() for x in SlotTextSplit]).intersection([x.lower() for x in self.legend])) :
                if list(set(modules_in_slot).intersection([x.lower() for x in self.legend])) :
                    #slot_modules = slotText.split('/')

                    for slot_module in slot_modules :

                        print(slot_module)

                        contrib_tmp = slot_module.replace(slot_module.split()[0] +' ', '') 
                        if len(contrib_tmp.split()) > 1 :
                            contrib_tmp = contrib_tmp.split()[0]

                        if (self.contributor in contrib_tmp) or  self.contributor == '' :

                            slot_module = slot_module.strip()

                            slot = CalendarSlot(    start_datetime=current_day_with_time,
                                                    ModuleName=slot_module.split()[0],
                                                    SlotText=slot_module
                                                )
                            fusion1 = False
                            fusion2 = False

                            # if slot.start_datetime.day == 27 and slot.start_datetime.month == 9 :
                            #     print(str(slot_modules) + "  -->  " + str(slot))

                            try:
                                fusion1 = CS.trySlotFusion( self.calendarSlots[-1], slot )
                                # if slot.start_datetime.day == 27 and slot.start_datetime.month == 9:
                                #     print(str(fusion1))

                            except IndexError:
                                print("")

                            try:
                                fusion2 = CS.trySlotFusion( self.calendarSlots[-2], slot )
                                # if slot.start_datetime.day == 27 and slot.start_datetime.month == 9:
                                #     print(str(fusion2))


                            except IndexError:
                                print("")

                            if not fusion1 and not fusion2 :
                                self.calendarSlots.append(slot)

                            #print ("\t %s" % slot )



    def printSlots(self):
        prev_date = self.calendarSlots[0].start_datetime
        for slot in self.calendarSlots :
            cur_date = slot.start_datetime

            if prev_date.year != cur_date.month and prev_date.month != cur_date.year and prev_date.day != cur_date.day :
                print("")
            prev_date = cur_date
            print("%s : %s   \t(%sh)" % (str(slot.start_datetime), slot.ModuleName, slot.timeLength[0]))


    def generateCsvGoogleCalendar(self):
        f = open('PlanningCPE.csv', 'w')
        f.write("Subject,Start Date,Start Time,End Date,End Time,All day event,Description,Location, Private\n")

        for slot in self.calendarSlots :

            start_date_str = "%02d/%02d/%d" % (slot.start_datetime.day, slot.start_datetime.month, slot.start_datetime.year)

            if slot.start_datetime.hour <= 12 :
                start_time_str = "%02d:%02d AM" % (slot.start_datetime.hour, slot.start_datetime.minute)
            else :
                start_time_str = "%02d:%02d PM" % (slot.start_datetime.hour-12, slot.start_datetime.minute)



            end_date_str = "%02d/%02d/%d" % (slot.start_datetime.day, slot.start_datetime.month, slot.start_datetime.year)

            end_time =  slot.start_datetime + timedelta(hours=slot.timeLength[0], minutes=slot.timeLength[1])
            if end_time.hour < 12:
                end_time_str = "%02d:%02d AM" % (end_time.hour, end_time.minute)
            elif end_time.hour > 12:
                end_time_str = "%02d:%02d PM" % (end_time.hour - 12, end_time.minute)
            else:
                end_time_str = "%02d:%02d PM" % (end_time.hour, end_time.minute)


            line = "%s,%s,%s,%s,%s,False,,%s,False" %  (
                slot.SlotText,          # Subject
                start_date_str,         # Start Date
                start_time_str,         # Start Time
                end_date_str,           # End Date
                end_time_str,            # End Time

                slot.ClassRoom

            )

            print(line)

            f.write(line + "\n")

        f.close()


# class GCalendarCpeGeneratorTest(unittest.TestCase):
#
#     def test_generateCalendarSlots(self):
#        self.assertTrue( True )


if __name__ == '__main__':
    #unittest.main()

    GCCG = GCalendarCpeGenerator()

    GCCG.getLegend()
    GCCG.generateCalendarSlots()
    GCCG.printSlots()
    GCCG.generateCsvGoogleCalendar()
